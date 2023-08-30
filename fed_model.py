#Fed Model
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl
import numpy as np
from datetime import datetime
path = "/Users/macbook/desktop/GF-FedModel/data.xlsx"

date_column = []
nominal_yield = []
earnings_yield = []
mmfData = []
mmf_yield=[]
name = "Fed Model by Day"
def readData(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(11,sheet.max_row +1):
        cell1 = sheet.cell(row = i, column = 2)
        cell2 = sheet.cell(row = i, column = 3)
        datecell = sheet.cell(row=i,column = 1)
        mmf = sheet.cell(row=i,column = 9)
        if i<2441:
            mmfData.insert(0,float(mmf.value))
        if (int(cell1.value)==0 or int(cell2.value)==0):
            continue
        else:
            nominal_yield.append(float(cell1.value))
            earnings_yield.append(100.0/(float(cell2.value)))
            date_column.append(datecell.value)

def readWeek(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(11,sheet.max_row +1,7):
        cell1 = sheet.cell(row = i, column = 2)
        cell2 = sheet.cell(row = i, column = 3)
        count = 0
        while (int(cell1.value)==0 or int(cell2.value)==0):
            count+=1
            cell1 = sheet.cell(row = i+count,column = 2)
            cell2 = sheet.cell(row = i+count,column = 3)
        
        nominal_yield.append(float(cell1.value))
        earnings_yield.append(100.0/(float(cell2.value)))
    name = "Fed Model by Week"

def plotData(x_data, yield1, yield2, name):
    plt.figure()
    plt.plot(x_data, yield1, label='Yu E Bao', color='blue')
    plt.plot(x_data, yield2, label='CSI 800', color='red')
    plt.xlabel('Year')
    plt.ylabel('Rate %')
    plt.title(name)
    plt.xticks(rotation=45)
    plt.legend()
    plt.show()

def plotResiduals(date_column,data,data1):
    plt.figure(figsize=(10, 6))
    a = mdates.date2num(date_column)
    # 使用numpy进行线性回归拟合
    coefficients = np.polyfit(a, data, 1)
    slope = coefficients[0]
    intercept = coefficients[1]
    trend_line = slope * mdates.date2num(date_column) + intercept

    # 计算残差（实际数据减去趋势线）
    residuals = data - trend_line

    coef = np.polyfit(a,data1,1)
    slope1 = coef[0]
    intercept1 = coef[1]
    trend_line1 = slope1 * mdates.date2num(date_column) + intercept1

    residuals1 = data1 - trend_line1
    # 绘制残差图
    plt.plot(date_column, residuals, label='Bond Yield', color='blue')
    plt.plot(date_column, residuals1, label='CSI 500 Earnings Yield', color='red')
    plt.axhline(y=0, color='green', linestyle='--', label='Zero line')
    plt.xlabel('Years')
    plt.ylabel('Residuals')
    plt.title('Residuals of Yield Data')
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    plt.gca().xaxis.set_major_locator(mdates.YearLocator())
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

readData(path)

#plotData(date_column, mmfData, earnings_yield, name)
plotResiduals(date_column,nominal_yield,earnings_yield)





